from uis.excel_checker import Ui_MainWindow
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
import collections
import openpyxl
import os
import sys
import time


class Core(QMainWindow, Ui_MainWindow):
	FOLDER_PATH = os.path.expanduser("~").replace("\\", "/") + "/Desktop"  # 默认文件路径

	def __init__(self):
		super(Core, self).__init__(None)  # 初始化
		self.setupUi(self)  # 启动
		self.setWindowTitle("excel基础校验工具")  # 设置标题

		'''绑定按钮事件'''
		self.pushButton_impt.clicked.connect(self.impt)
		self.pushButton_check.clicked.connect(self.check)
		self.pushButton_expt.clicked.connect(self.expt)

		'''设置只读'''
		self.lineEdit_impt.setReadOnly(True)
		self.textEdit_report.setReadOnly(True)

		'''取消自动换行'''
		self.textEdit_report.setLineWrapMode(False)

		'''全局变量'''
		self.workbook = None  # 当前工作簿
		self.worksheet = None  # 当前工作表
		self.titles = None  # 标题集合
		self.report = None  # 校验报告

	def impt(self):  # 导入文件
		filepath = QFileDialog.getOpenFileName(
			self, filter="*.xlsx", directory=self.FOLDER_PATH, options=QFileDialog.Options()
		)[0].strip()  # 文件对话框

		if filepath:
			self.lineEdit_impt.setText(filepath)
			if self.workbook:
				self.workbook.close()  # 重新导入时关闭前一工作簿的IO流
			self.workbook = openpyxl.load_workbook(filepath, read_only=True)  # 工作簿（只读模式）
			self.worksheet = self.workbook.active  # 当前工作表

			'''读取标题'''
			self.titles = [str(title.value) for title in list(
				self.worksheet.iter_rows(min_col=1, min_row=1, max_row=1)
			)[0] if title.value]  # 标题集合
			self.comboBox_empty.set_items(self.titles)  # 空值下拉框
			self.comboBox_repeat.set_items(self.titles)  # 重复值下拉框

	def check(self):  # 校验数据
		if self.workbook:
			columns_empty = self.comboBox_empty.get_items()  # 校验空值列
			columns_repeat = self.comboBox_repeat.get_items()  # 校验重复值列
			idxs_empty = {i: title for i, title in enumerate(self.titles) if title in columns_empty}  # 空值列索引
			idxs_repeat = {i: title for i, title in enumerate(self.titles) if title in columns_repeat}  # 重复值列索引

			if idxs_empty or idxs_repeat:
				dic_empty = collections.defaultdict(list)  # 校验空值列字典
				dic_repeat_initial = collections.defaultdict(dict)  # 校验重复值列字典（首次出现）
				dic_repeat_noninitial = collections.defaultdict(dict)  # 校验重复值列字典（非首次出现）

				'''遍历导入文件'''
				for i, data_row in enumerate(
						self.worksheet.iter_rows(min_col=1, max_col=self.worksheet.max_column, min_row=2)):
					for j, data_cell in enumerate(data_row):
						if j in idxs_empty:
							if not data_cell.value:  # 空值
								dic_empty[j].append(i + 2)  # 添加excel行号
						if j in idxs_repeat:
							if data_cell.value:  # 非空值
								if data_cell.value in dic_repeat_initial[j]:  # 非首次出现
									if data_cell.value in dic_repeat_noninitial[j]:
										dic_repeat_noninitial[j][data_cell.value].append(i + 2)  # 添加excel行号
									else:
										dic_repeat_noninitial[j][data_cell.value] = [i + 2]  # 添加excel行号
								else:  # 首次出现
									dic_repeat_initial[j][data_cell.value] = i + 2  # 添加excel行号

				'''添加重复值首次出现的excel行号'''
				if idxs_repeat:
					for repeat_column in dic_repeat_noninitial:
						for repeat_row in dic_repeat_noninitial[repeat_column]:
							dic_repeat_noninitial[repeat_column][repeat_row].insert(
								0, dic_repeat_initial[repeat_column][repeat_row])  # 添加首次出现的excel行号

				'''生成校验空值报告'''
				report_empty = ""
				if not idxs_empty:
					report_empty = "无校验空值列"
				elif not dic_empty:
					report_empty = "无空值数据"
				else:
					for empty_column in dic_empty:
						report_empty += f"列【{idxs_empty[empty_column]}】空值行号: "
						for empty_row in dic_empty[empty_column]:
							report_empty += f"{empty_row}, "
						report_empty = report_empty.strip(", ") + "\n\n"
					report_empty = report_empty.strip(", ").strip()

				'''生成校验重复值报告'''
				report_repeat = ""
				if not idxs_repeat:
					report_repeat = "无校验重复值列"
				elif not dic_repeat_noninitial:
					report_repeat = "无重复值数据"
				else:
					for repeat_column in dic_repeat_noninitial:
						for repeat_value in dic_repeat_noninitial[repeat_column]:
							report_repeat += f"列【{idxs_repeat[repeat_column]}】值【{repeat_value}】重复行号: "
							for repeat_row in dic_repeat_noninitial[repeat_column][repeat_value]:
								report_repeat += f"{repeat_row}, "
							report_repeat = report_repeat.strip(", ") + "\n\n"
					report_repeat = report_repeat.strip(", ").strip()

				self.report = report_empty + f"\n\n{'-' * 64}\n\n" + report_repeat  # 校验报告
				self.textEdit_report.setText(self.report)

	def expt(self):
		if self.report:
			current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())  # 当前时间
			filename = f"数据基础校验报告{time.strftime('%Y%m%d%H%M%S', time.localtime())}"  # 默认文件名

			filepath = QFileDialog.getSaveFileName(
				self, filter="*.txt", directory=f"{self.FOLDER_PATH}/{filename}",
				options=QFileDialog.Options())[0].strip()  # 文件对话框
			if filepath:
				with open(filepath, mode="w", encoding="utf-8") as file:
					file.write(f"-------- 报告生成时间 [{current_time}] --------\n\n{self.report}")  # 写入文件


if __name__ == "__main__":
	QCoreApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)  # 自动适应高分辨率
	app = QApplication(sys.argv)  # 创建app（sys.argv是命令行参数列表）
	myWindow = Core()  # 初始化
	myWindow.show()  # 显示窗口控件
	sys.exit(app.exec_())  # 完整退出
